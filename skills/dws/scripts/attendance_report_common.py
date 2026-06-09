#!/usr/bin/env python3
"""
考勤报表导出脚本 — 公共模块

[AI Agent 强制门禁] 本模块不可单独执行，且任何调用方脚本
   （attendance_report_detail/monthly/daily.py）执行前都必须先阅读：
   dingtalk-workspace/references/products/attendance-report.md

   工作流细节、报表类型判断、人员获取、列选择、错误处理等约束
   全部在 attendance-report.md，禁止凭本脚本源码或 --help 自行组装命令。

被 attendance_report_detail.py / attendance_report_monthly.py /
attendance_report_daily.py 三个粒度脚本共享。

职责：
  1. dws CLI 调用（run_dws / run_dws_raw）
  2. 接口分批 / 切片（chunk_users / slice_date_range）
  3. dws 返回值通用解析（unwrap_result / extract_records）
  4. 字段（columns）模糊匹配（match_columns_by_keywords）
  5. userId → name 映射（resolve_user_names）
  6. Excel 写入（write_excel）
  7. 错误处理 / stderr 进度日志

约束：
  - 不依赖 dws 命令的具体业务字段（除接口顶层 success/result/error）
  - 业务字段解析全部由各粒度脚本负责
  - 时间字段单位由各粒度脚本自行处理（attendance 接口多为毫秒时间戳）
"""

from __future__ import annotations

import hashlib
import json
import os
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Any, Iterable

# ─────────────────────────────────────────────────────────────────────────────
# 常量：dws 接口限制（来自 attendance.md）
# ─────────────────────────────────────────────────────────────────────────────

MAX_USERS_PER_BATCH = 5             # report query-data: --users 最多 5 人（降低单批人数避免超时）
MAX_DAYS_PER_SLICE = 32             # report query-data: --start 到 --end ≤ 32 天
DWS_TIMEOUT_SECONDS = 120           # 单次 dws 调用超时
DATETIME_FMT = "%Y-%m-%d %H:%M:%S"
DATE_FMT = "%Y-%m-%d"


# ─────────────────────────────────────────────────────────────────────────────
# stderr 日志（脚本所有进度信息打 stderr，stdout 留给最终摘要）
# ─────────────────────────────────────────────────────────────────────────────

def log(msg: str) -> None:
    """打印进度信息到 stderr，stdout 保留给最终摘要。"""
    print(msg, file=sys.stderr, flush=True)


def warn(msg: str) -> None:
    log(f"[WARN] {msg}")


def error(msg: str) -> None:
    log(f"[ERROR] {msg}")


# ─────────────────────────────────────────────────────────────────────────────
# dws 调用
# ─────────────────────────────────────────────────────────────────────────────

class DwsCallError(Exception):
    """dws 调用失败（含进程退出非零、超时、JSON 解析失败、业务 success=false）。"""

    def __init__(self, message: str, *, is_permission_error: bool = False) -> None:
        super().__init__(message)
        self.is_permission_error = is_permission_error


def _looks_like_permission_error(text: str) -> bool:
    """启发式：判断错误文本是否属于权限/管理员问题。"""
    if not text:
        return False
    lower = text.lower()
    keywords = ("403", "permission", "denied", "unauthorized",
                "forbidden", "无权限", "权限不足", "管理员")
    return any(k in lower for k in keywords)


def run_dws(args: list[str]) -> Any:
    """
    调用 `dws <args> --format json`，返回解析后的 JSON。

    自动追加 `--format json`（如果调用方没传），并解开顶层 `success/result/error`：
      - success=True  → 返回 result 内容
      - success=False → 抛 DwsCallError（含 is_permission_error 标记）
      - 进程退出非零或解析失败 → 抛 DwsCallError

    注意：本函数仅做"顶层解包"，业务字段解析由调用方负责。
    """
    if "--format" not in args:
        args = args + ["--format", "json"]

    cmd = ["dws"] + args
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=DWS_TIMEOUT_SECONDS,
        )
    except subprocess.TimeoutExpired as e:
        raise DwsCallError(f"dws 调用超时（{DWS_TIMEOUT_SECONDS}s）：{' '.join(cmd)}") from e
    except FileNotFoundError as e:
        raise DwsCallError("未找到 dws 命令，请确认 dws CLI 已安装并在 PATH 中") from e

    stdout = result.stdout or ""
    stderr = result.stderr or ""

    if result.returncode != 0:
        is_perm = _looks_like_permission_error(stderr) or _looks_like_permission_error(stdout)
        raise DwsCallError(
            f"dws 调用失败（exit={result.returncode}）: {stderr.strip() or stdout.strip()}",
            is_permission_error=is_perm,
        )

    try:
        data = json.loads(stdout)
    except json.JSONDecodeError as e:
        raise DwsCallError(f"dws 返回非 JSON：{stdout[:200]!r}") from e

    return unwrap_result(data)


def unwrap_result(data: Any) -> Any:
    """
    解开 dws 返回的顶层 `{success, result, error}` 包装。

    success=True  → 返回 result（可能是 dict / list / None）
    success=False → 抛 DwsCallError
    其他形态     → 原样返回（兼容部分接口直接返回数据）
    """
    if not isinstance(data, dict):
        return data

    if "success" not in data:
        # 不是标准包装，原样返回
        return data

    if data.get("success") is True:
        return data.get("result")

    # success = False
    err = data.get("error") or {}
    if isinstance(err, dict):
        msg = err.get("message") or err.get("msg") or json.dumps(err, ensure_ascii=False)
    else:
        msg = str(err)
    raise DwsCallError(
        f"dws 业务失败：{msg}",
        is_permission_error=_looks_like_permission_error(msg),
    )


# ─────────────────────────────────────────────────────────────────────────────
# 通用记录提取（兼容多种数据嵌套形态）
# ─────────────────────────────────────────────────────────────────────────────

def extract_records(payload: Any) -> list[dict]:
    """
    从 dws 返回的 result 中提取"记录数组"。

    兼容多种常见嵌套：
      - 直接是 list[dict]                → 原样返回
      - {"data": [...]}                  → 取 data
      - {"records": [...]}               → 取 records
      - {"list": [...]}                  → 取 list
      - {"items": [...]}                 → 取 items
      - {"result": [...]} (双层包装)     → 递归一次
      - 其他 dict 但只有一个值是 list    → 取那个 list
      - 其他形态                         → 返回 []，并 warn

    业务字段不在本函数关心范围内。
    """
    if payload is None:
        return []
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("data", "records", "list", "items", "result"):
            if key in payload and isinstance(payload[key], list):
                return [item for item in payload[key] if isinstance(item, dict)]
        # 兜底：dict 中只有一个 list 值
        list_values = [v for v in payload.values() if isinstance(v, list)]
        if len(list_values) == 1:
            return [item for item in list_values[0] if isinstance(item, dict)]
        warn(f"未能从返回中识别记录数组，顶层 keys={list(payload.keys())}")
        return []
    warn(f"未能识别返回类型：{type(payload).__name__}")
    return []


def flatten_query_data_records(
    records: list[dict],
    column_id_to_name: dict[str, str] | None = None,
) -> list[dict]:
    """
    展平 report query-data 返回的嵌套 values 结构。

    接口原始格式：
      {"userId":"xxx", "values":[{"termId":"173410778","value":"1"}, ...], "workDate":"2026-05-01"}

    展平后：
      {"userId":"xxx", "workDate":"2026-05-01", "173410778":"1", "节假日+出勤":"1", ...}

    Args:
        records: extract_records 返回的原始记录列表
        column_id_to_name: 可选的 columnId → columnName 映射，展平时同时写入字段名 key
    """
    flattened: list[dict] = []
    for record in records:
        values_list = record.get("values")
        if not isinstance(values_list, list):
            # 已经是平铺格式或无 values 字段，原样保留
            flattened.append(record)
            continue
        flat: dict[str, Any] = {}
        # 保留顶层非 values 字段（userId, workDate, corpId 等）
        for k, v in record.items():
            if k != "values":
                flat[k] = v
        # 展平 values 数组
        for entry in values_list:
            if not isinstance(entry, dict):
                continue
            term_id = str(entry.get("termId", entry.get("columnId", entry.get("id", ""))))
            value = entry.get("value", entry.get("data", ""))
            if term_id:
                flat[term_id] = value
                # 同时写入字段名 key（方便按名称取值）
                if column_id_to_name and term_id in column_id_to_name:
                    flat[column_id_to_name[term_id]] = value
        flattened.append(flat)
    return flattened


def extract_group_names_from_records(
    records: list[dict],
    user_ids: list[str],
) -> dict[str, str]:
    """
    获取每个用户的考勤组名称。

    优先从 report query-data 原始记录中提取（如果接口返回了 groupName 字段），
    否则回退到通过 `dws attendance group search` + `filtered-get --member`
    获取所有考勤组的成员列表，反向映射 userId → 考勤组名称。

    返回 {userId: groupName} 映射，未找到的用户映射为空字符串。
    """
    group_map: dict[str, str] = {}
    candidate_keys = ("groupName", "group_name", "attendanceGroupName",
                      "groupId", "group_id")

    # 1) 先尝试从原始记录中提取
    for record in records:
        uid = _first_nonempty(record, ("userId", "userid", "user_id", "targetUserId"))
        if uid is None:
            continue
        uid_str = str(uid)
        if uid_str in group_map:
            continue
        name = _first_nonempty(record, candidate_keys)
        if name is not None and str(name).strip():
            group_map[uid_str] = str(name).strip()

    # 2) 如果还有用户未匹配到考勤组，通过 group API 反向查找
    missing_uids = {uid for uid in user_ids if uid not in group_map}
    if missing_uids:
        api_map = _resolve_group_names_via_api(missing_uids)
        group_map.update(api_map)

    # 兜底：未找到的用户填空字符串
    for uid in user_ids:
        if uid not in group_map:
            group_map[uid] = ""

    return group_map


def _resolve_group_names_via_api(target_uids: set[str]) -> dict[str, str]:
    """
    通过 dws attendance group search + filtered-get --member
    反向映射 userId → 考勤组名称。

    流程：
      1. group search 获取所有考勤组（id + name）
      2. 对每个考勤组调用 filtered-get --member 获取成员 userId 列表
      3. 将 target_uids 中的用户与考勤组成员做交集映射
    """
    result_map: dict[str, str] = {}
    if not target_uids:
        return result_map

    # 获取所有考勤组
    try:
        search_payload = run_dws([
            "attendance", "group", "search",
            "--limit", "200",
        ])
    except DwsCallError as e:
        log(f"[group] 获取考勤组列表失败：{e}")
        return result_map

    search_result = unwrap_result(search_payload)
    groups: list[dict] = []
    if isinstance(search_result, dict):
        groups = search_result.get("items", [])
    elif isinstance(search_result, list):
        groups = search_result

    if not groups:
        log("[group] 未获取到任何考勤组")
        return result_map

    log(f"[group] 共 {len(groups)} 个考勤组，开始查询成员列表")

    remaining = set(target_uids)
    for group in groups:
        if not remaining:
            break
        group_id = group.get("id")
        group_name = group.get("name", "")
        if not group_id:
            continue

        try:
            detail_payload = run_dws([
                "attendance", "group", "filtered-get",
                "--group-id", str(group_id),
                "--member",
            ])
        except DwsCallError as e:
            log(f"[group] filtered-get 失败 (group={group_name}): {e}")
            continue

        detail = unwrap_result(detail_payload)
        if not isinstance(detail, dict):
            continue

        member_users = detail.get("memberUsers", [])
        if not isinstance(member_users, list):
            continue

        for member_uid in member_users:
            uid_str = str(member_uid)
            if uid_str in remaining:
                result_map[uid_str] = group_name
                remaining.discard(uid_str)

    if remaining:
        log(f"[group] {len(remaining)} 个用户未匹配到考勤组")

    return result_map


def dump_first_record_for_inspection(records: list[dict], label: str) -> None:
    """
    第一次跑脚本时，把第一条记录打到 stderr，方便用户/开发者
    看清真实字段结构后回来调优解析逻辑。
    """
    if not records:
        log(f"[inspect:{label}] 无记录")
        return
    sample = records[0]
    log(f"[inspect:{label}] 首条记录字段示例（用于核对真实结构）：")
    log(json.dumps(sample, ensure_ascii=False, indent=2))


# ─────────────────────────────────────────────────────────────────────────────
# 时间区间切片
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class DateSlice:
    start: datetime  # 含
    end: datetime    # 含

    @property
    def start_str(self) -> str:
        return self.start.strftime(DATETIME_FMT)

    @property
    def end_str(self) -> str:
        return self.end.strftime(DATETIME_FMT)

    @property
    def label(self) -> str:
        return f"{self.start.strftime(DATE_FMT)}~{self.end.strftime(DATE_FMT)}"


def parse_datetime_arg(s: str, *, end_of_day: bool = False) -> datetime:
    """
    解析用户输入的日期参数，支持：
      - YYYY-MM-DD                  → 00:00:00 或 23:59:59（取决于 end_of_day）
      - YYYY-MM-DD HH:mm:ss
    """
    s = s.strip()
    try:
        return datetime.strptime(s, DATETIME_FMT)
    except ValueError:
        pass
    try:
        d = datetime.strptime(s, DATE_FMT)
        if end_of_day:
            return d.replace(hour=23, minute=59, second=59)
        return d
    except ValueError as e:
        raise ValueError(
            f"无法解析日期 {s!r}，请使用 YYYY-MM-DD 或 YYYY-MM-DD HH:mm:ss"
        ) from e


def slice_date_range(
    start: datetime,
    end: datetime,
    max_days: int = MAX_DAYS_PER_SLICE,
) -> list[DateSlice]:
    """
    把 [start, end] 切成多个 ≤ max_days 天的小区间。

    每片含起含止；最后一片可能短于 max_days。
    """
    if end < start:
        raise ValueError(f"结束时间 {end} 早于开始时间 {start}")

    slices: list[DateSlice] = []
    cur_start = start
    while cur_start <= end:
        # 这片的最晚结束时间（不超 max_days，且不超 end）
        cur_end_limit = cur_start + timedelta(days=max_days - 1)
        cur_end_limit = cur_end_limit.replace(hour=23, minute=59, second=59)
        cur_end = min(cur_end_limit, end)
        slices.append(DateSlice(start=cur_start, end=cur_end))
        # 下一片从次日 00:00 开始
        next_day = (cur_end + timedelta(seconds=1)).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        cur_start = next_day
    return slices


# ─────────────────────────────────────────────────────────────────────────────
# 用户分批
# ─────────────────────────────────────────────────────────────────────────────

def chunk_users(users: list[str], size: int = MAX_USERS_PER_BATCH) -> list[list[str]]:
    """把 userId 列表切成每片 ≤ size 的小批。"""
    if size <= 0:
        raise ValueError(f"size 必须 > 0，得到 {size}")
    return [users[i: i + size] for i in range(0, len(users), size)]


# ─────────────────────────────────────────────────────────────────────────────
# columns 模糊匹配
# ─────────────────────────────────────────────────────────────────────────────

def match_columns_by_keywords(
    all_columns: list[dict],
    keywords: list[str],
    *,
    name_keys: tuple[str, ...] = ("name", "columnName", "title", "label"),
    id_keys: tuple[str, ...] = ("id", "columnId", "code", "key"),
) -> list[dict]:
    """
    在 report columns 返回的字段列表里，按关键词匹配目标字段。

    匹配策略（按优先级）：
      1. 精确匹配：关键词 == 字段名（优先）
      2. 子串匹配：关键词是字段名的子串（仅当精确匹配无结果时回退）

    列名严格使用接口返回的原始字段名，不做任何修改。
    """
    matched: list[dict] = []
    matched_ids: set[str] = set()
    hit_keywords: set[str] = set()

    # 构建 name → (col_dict, cid_str) 索引
    col_index: list[tuple[str, str, dict]] = []
    for col in all_columns:
        name = _first_nonempty(col, name_keys)
        cid = _first_nonempty(col, id_keys)
        if not name or cid is None:
            continue
        col_index.append((str(name), str(cid), col))

    for kw in keywords:
        kw_stripped = kw.strip()
        if not kw_stripped:
            continue

        # 第一轮：精确匹配
        exact_hit = False
        for name, cid_str, col in col_index:
            if name == kw_stripped and cid_str not in matched_ids:
                enriched = dict(col)
                enriched["_column_id"] = cid_str
                enriched["_column_name"] = name
                matched.append(enriched)
                matched_ids.add(cid_str)
                hit_keywords.add(kw_stripped)
                exact_hit = True
                break

        if exact_hit:
            continue

        # 第二轮：子串匹配（回退），只取第一个命中
        kw_lower = kw_stripped.lower()
        for name, cid_str, col in col_index:
            if kw_lower in name.lower() and cid_str not in matched_ids:
                enriched = dict(col)
                enriched["_column_id"] = cid_str
                enriched["_column_name"] = name
                matched.append(enriched)
                matched_ids.add(cid_str)
                hit_keywords.add(kw_stripped)
                break

    missing = set(k.strip() for k in keywords if k.strip()) - hit_keywords
    if missing:
        warn(f"以下关键词未匹配到任何字段，已跳过：{sorted(missing)}")
    return matched


def _first_nonempty(d: dict, keys: Iterable[str]) -> Any:
    for k in keys:
        if k in d and d[k] not in (None, ""):
            return d[k]
    return None


# ─────────────────────────────────────────────────────────────────────────────
# userId → name 映射
# ─────────────────────────────────────────────────────────────────────────────

def resolve_users_from_input(raw_ids: list[str]) -> list[str]:
    """
    智能解析 --users 输入：自动区分部门ID和员工userId。

    Wukong Agent 经常把部门ID当 userId 传入。本函数尝试对每个ID调用
    `dws contact dept list-members` 获取成员列表：
      - 如果成功且返回了员工，说明该ID是部门ID，展开为员工userId列表
      - 如果失败或无结果，说明该ID本身就是userId，原样保留

    最终返回去重后的 userId 列表。
    """
    if not raw_ids:
        return []

    resolved: list[str] = []
    seen: set[str] = set()

    # 先尝试批量查部门成员（可能全是部门ID）
    try:
        result = run_dws([
            "contact", "dept", "list-members",
            "--ids", ",".join(raw_ids),
        ])
        members = extract_records(result)
        if members:
            # 成功获取到成员 → 输入是部门ID
            for member in members:
                uid = _first_nonempty(member, ("userId", "userid", "id"))
                if uid and str(uid) not in seen:
                    resolved.append(str(uid))
                    seen.add(str(uid))
            if resolved:
                log(f"[users] 检测到输入为部门ID，已展开为 {len(resolved)} 个员工userId")
                return resolved
    except DwsCallError:
        pass

    # 逐个ID尝试：可能混合了部门ID和userId
    for raw_id in raw_ids:
        if raw_id in seen:
            continue
        try:
            result = run_dws([
                "contact", "dept", "list-members",
                "--ids", raw_id,
            ])
            members = extract_records(result)
            if members:
                for member in members:
                    uid = _first_nonempty(member, ("userId", "userid", "id"))
                    if uid and str(uid) not in seen:
                        resolved.append(str(uid))
                        seen.add(str(uid))
                log(f"[users] 部门ID {raw_id} 展开为 {len(members)} 个员工")
                continue
        except DwsCallError:
            pass
        # 不是部门ID，当作userId保留
        if raw_id not in seen:
            resolved.append(raw_id)
            seen.add(raw_id)

    return resolved


@dataclass
class UserInfo:
    """用户基础信息，用于报表的姓名/部门/工号/职位列。"""
    name: str = ""
    dept_name: str = ""
    job_number: str = ""
    title: str = ""


def _extract_title_from_labels(labels: list) -> str:
    """从 orgEmployeeModel.labels 数组中提取职务名称。"""
    if not isinstance(labels, list):
        return ""
    for item in labels:
        if isinstance(item, dict) and item.get("groupName") == "职务":
            name = item.get("name", "")
            if name:
                return str(name)
    return ""


def _parse_user_record(record: dict) -> tuple[str, UserInfo] | None:
    """
    从 dws contact user get 返回的单条记录中解析用户信息。

    接口返回结构为嵌套格式：
      {"orgEmployeeModel": {"userId": "xxx", "orgUserName": "吾贤",
       "depts": [{"deptName": "技术部"}],
       "labels": [{"groupName": "职务", "name": "财务"}],
       ...}, "isAdmin": true}

    也兼容扁平格式（其他接口可能返回）：
      {"userId": "xxx", "name": "吾贤", "deptName": "技术部", ...}
    """
    # 优先从嵌套的 orgEmployeeModel 中提取
    model = record.get("orgEmployeeModel")
    if isinstance(model, dict):
        uid = model.get("userId") or model.get("orgUserId")
        if not uid:
            return None
        name = model.get("orgUserName") or model.get("name") or ""
        depts = model.get("depts") or []
        dept_name = depts[0].get("deptName", "") if depts and isinstance(depts[0], dict) else ""
        # 工号：尝试多个候选字段
        job_number = (model.get("jobNumber") or model.get("workNumber")
                      or model.get("empId") or "")
        # 职位：优先 title/position，回退到 labels 中 groupName=="职务" 的条目
        title = (model.get("title") or model.get("position")
                 or _extract_title_from_labels(model.get("labels", [])))
        return str(uid), UserInfo(
            name=str(name),
            dept_name=str(dept_name),
            job_number=str(job_number),
            title=str(title),
        )

    # 回退：扁平格式
    uid = _first_nonempty(record, ("userId", "userid", "id"))
    if not uid:
        return None
    return str(uid), UserInfo(
        name=str(_first_nonempty(record, ("name", "userName", "nick")) or ""),
        dept_name=str(_first_nonempty(record, ("deptName", "dept_name", "department")) or ""),
        job_number=str(_first_nonempty(record, ("jobNumber", "job_number", "workNumber")) or ""),
        title=str(_first_nonempty(record, ("title", "position", "jobTitle")) or ""),
    )


def resolve_user_info(user_ids: list[str]) -> dict[str, UserInfo]:
    """
    批量获取用户的完整基础信息（姓名、部门、工号、职位）。

    适配 dws contact user get 返回的嵌套 orgEmployeeModel 结构。
    分批处理（每批20人），失败时不抛错。
    """
    if not user_ids:
        return {}

    info_map: dict[str, UserInfo] = {}

    for i in range(0, len(user_ids), 20):
        batch = user_ids[i:i + 20]
        try:
            result = run_dws(["contact", "user", "get", "--ids", ",".join(batch)])
            for record in extract_records(result):
                parsed = _parse_user_record(record)
                if parsed:
                    uid_str, info = parsed
                    info_map[uid_str] = info
        except DwsCallError as e:
            log(f"[user-info] 批量 get 失败（batch {i // 20 + 1}）：{e}")

    # 兜底：未解析到的用户填充 userId 作为姓名
    for uid in user_ids:
        if uid not in info_map:
            info_map[uid] = UserInfo(name=uid)

    return info_map


def resolve_user_names(user_ids: list[str]) -> dict[str, str]:
    """
    给一组 userId 解析姓名映射（向后兼容接口）。

    内部调用 resolve_user_info，只返回 {userId: name} 映射。
    """
    info_map = resolve_user_info(user_ids)
    return {uid: info.name or uid for uid, info in info_map.items()}


# ─────────────────────────────────────────────────────────────────────────────
# Excel 写入
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# Excel 样式常量与辅助函数
# ─────────────────────────────────────────────────────────────────────────────

# 配色（参考钉钉考勤报表风格：青绿标题 + 浅黄表头 + 白色数据区）
_TITLE_FILL_COLOR = "D5EAEA"       # 浅青绿 — 主标题背景
_TITLE_FONT_COLOR = "1F6E6E"       # 深青 — 主标题字体
_SUBTITLE_FILL_COLOR = "DAEEF3"    # 浅蓝 — 副标题（生成时间）背景
_SUBTITLE_FONT_COLOR = "31708F"    # 深蓝 — 副标题字体

_HEADER_FILL_COLOR = "FFF2CC"      # 浅黄 — 表头背景
_HEADER_FONT_COLOR = "333333"      # 深灰近黑 — 表头字体
_DATA_FONT_COLOR = "333333"        # 数据字体颜色

_BORDER_COLOR = "BFBFBF"           # 浅灰 — 单元格边框
_FONT_NAME = "微软雅黑"

# 日历表人员交替配色（奇数人白底，偶数人浅灰蓝底，便于区分不同人员）
_CALENDAR_BAND_COLORS = ("FFFFFF", "EDF2F9")

# 考勤结果单元格条件配色（参考钉钉 previewStyleByValue 配置）
# 规则按优先级排列，首个匹配命中即停止；无 color 键表示不填充背景色。
import re as _re

_ATTEND_RESULT_STYLE_RULES: list[tuple["_re.Pattern[str]", str | None]] = [
    # 白底（红字加粗由字体控制）：周末 — POI index 9 WHITE (255,255,255)
    (_re.compile(r".*(星期六|星期日|星期天)[\s\S]*"), "FFFFFF"),
    # 浅黄(TAN)：补卡审批通过/举证打卡/加班/外出/假/调休 — POI index 47 (255,204,153)
    (_re.compile(r".*(补卡审批通过|举证打卡审批通过|加班|外出|假|调休)[\s\S]*"), "FFCC99"),
    # 浅青(LIGHT_TURQUOISE)：外勤/出差 — POI index 41 (204,255,255)
    (_re.compile(r".*(外勤|出差)[\s\S]*"), "CCFFFF"),
    # 浅黄(TAN)：管理员改为正常 — POI index 47 (255,204,153)
    (_re.compile(r"(?=.*管理员)(?=.*改为正常)^[\s\S]*$"), "FFCC99"),
    # 水蓝(AQUA)：旷工迟到（须在"旷工""迟到"之前匹配）— POI index 49 (51,204,204)
    (_re.compile(r".*(旷工迟到)[\s\S]*"), "33CCCC"),
    # 玫红(ROSE)：旷工 — POI index 45 (255,153,204)
    (_re.compile(r".*(旷工)[\s\S]*"), "FF99CC"),
    # 淡蓝(PALE_BLUE)：严重迟到（须在普通"迟到"之前匹配）— POI index 44 (153,204,255)
    (_re.compile(r".*(严重迟到)[\s\S]*"), "99CCFF"),
    # 浅绿(LIGHT_GREEN)：迟到 — POI index 42 (204,255,204)
    (_re.compile(r".*(迟到)[\s\S]*"), "CCFFCC"),
    # 柠檬黄(LEMON_CHIFFON)：早退 — POI index 26 (255,255,153)
    (_re.compile(r".*(早退)[\s\S]*"), "FFFF99"),
    # 珊瑚粉(CORAL)：缺卡 — POI index 29 (255,128,128)
    (_re.compile(r".*(缺卡)[\s\S]*"), "FF8080"),
    # 无色：未排班/休息/正常 — 不填充
    (_re.compile(r".*(未排班|休息|正常)[\s\S]*"), None),
]


def _match_attend_result_color(text: str) -> str | None:
    """
    根据考勤结果文本匹配颜色（6 位 hex，不含 #），无匹配或匹配到"无色"规则时返回 None。
    """
    if not text:
        return None
    for pattern, color in _ATTEND_RESULT_STYLE_RULES:
        if pattern.fullmatch(text):
            return color
    return None

# 行高
_TITLE_ROW_HEIGHT = 30             # 主标题行高
_SUBTITLE_ROW_HEIGHT = 22          # 副标题行高
_HEADER_ROW_HEIGHT = 32            # 表头行高（更高，配合浅黄底色）
_DATA_ROW_HEIGHT = 22              # 数据行高

# 列宽估算：中文字符权重，宽度上下限
_CJK_CHAR_WEIGHT = 2.0
_ASCII_CHAR_WEIGHT = 1.1
_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 40


def _is_cjk(ch: str) -> bool:
    """判断一个字符是否为中日韩字符（用于估算 Excel 列宽）。"""
    if not ch:
        return False
    code = ord(ch)
    return (
        0x4E00 <= code <= 0x9FFF       # CJK 统一汉字
        or 0x3000 <= code <= 0x303F    # CJK 符号和标点
        or 0xFF00 <= code <= 0xFFEF    # 全角字符
    )


def _estimate_text_width(text: str) -> float:
    """估算字符串在 Excel 中显示所占的列宽（中文 ≈ 2，ASCII ≈ 1）。"""
    if not text:
        return 0.0
    width = 0.0
    for ch in text:
        width += _CJK_CHAR_WEIGHT if _is_cjk(ch) else _ASCII_CHAR_WEIGHT
    return width


def _is_numeric_value(value: Any) -> bool:
    """判断一个值是否为可对齐到右侧的数字（或纯数字字符串）。"""
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return False
        try:
            float(s)
            return True
        except ValueError:
            return False
    return False


def _build_styles():
    """构建 Excel 样式对象集合，避免每个单元格重复创建。"""
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )

    thin_side = Side(border_style="thin", color=_BORDER_COLOR)
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    return {
        # 主标题（"月度汇总展示  统计日期：xxx 至 xxx"）
        "title_font": Font(name=_FONT_NAME, bold=True, color=_TITLE_FONT_COLOR, size=14),
        "title_fill": PatternFill(fill_type="solid", fgColor=_TITLE_FILL_COLOR),
        "title_align": Alignment(horizontal="left", vertical="center", indent=1),
        # 副标题（"报表生成时间：xxx"）
        "subtitle_font": Font(name=_FONT_NAME, color=_SUBTITLE_FONT_COLOR, size=10),
        "subtitle_fill": PatternFill(fill_type="solid", fgColor=_SUBTITLE_FILL_COLOR),
        "subtitle_align": Alignment(horizontal="left", vertical="center", indent=1),
        # 表头
        "header_font": Font(name=_FONT_NAME, bold=True, color=_HEADER_FONT_COLOR, size=11),
        "header_fill": PatternFill(fill_type="solid", fgColor=_HEADER_FILL_COLOR),
        "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        # 数据
        "data_font": Font(name=_FONT_NAME, color=_DATA_FONT_COLOR, size=10),
        "align_left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "align_right": Alignment(horizontal="right", vertical="center", wrap_text=False),
        "align_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": border,
        # 日历表人员交替配色（按 merge_groups 块交替）
        "band_fills": tuple(
            PatternFill(fill_type="solid", fgColor=c)
            for c in _CALENDAR_BAND_COLORS
        ),
    }


def _apply_sheet_styles(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    styles: dict,
    *,
    title: str | None = None,
    subtitle: str | None = None,
    freeze_first_col: bool = True,
    merge_groups: list[tuple[int, int, int]] | None = None,
    attend_result_columns: set[int] | None = None,
    attend_result_rows: set[int] | None = None,
    image_columns: list[str] | None = None,
    image_size: tuple[int, int] = (80, 120),
) -> None:
    """
    给一个已写入数据的 sheet 应用统一样式（钉钉风格）。

    布局（自上而下）：
      [可选] 第 1 行：主标题（青绿底，跨所有列，写"xxx展示  统计日期：A 至 B"）
      [可选] 第 2 行：副标题（浅蓝底，跨所有列，写"报表生成时间：xxx"）
      表头行：浅黄底 + 加粗黑字 + 居中 + 加高
      数据行：白底 + 居中 + 灰色细边框

    参数:
      title:    可选主标题文本（如"月度汇总展示  统计日期：2026-01-01 至 2026-01-31"）
      subtitle: 可选副标题文本（如"报表生成时间：2026-01-20 15:46"）
      freeze_first_col: 是否冻结首列（姓名）
      attend_result_columns: 考勤结果列的 0-based 列索引集合（月度汇总场景）
      attend_result_rows:    考勤结果行的 0-based 行偏移集合（日历表场景）
    """
    from openpyxl.utils import get_column_letter

    n_cols = len(headers)
    n_rows = len(rows)
    last_col_letter = get_column_letter(n_cols) if n_cols >= 1 else "A"

    # ── 标题区（如果有）── 占据 1~2 行，跨所有列 ───────────────────────
    title_row_count = 0
    if title:
        title_row_count += 1
        title_row = title_row_count
        ws.cell(row=title_row, column=1, value=title)
        if n_cols >= 2:
            ws.merge_cells(
                start_row=title_row, start_column=1,
                end_row=title_row, end_column=n_cols,
            )
        cell = ws.cell(row=title_row, column=1)
        cell.font = styles["title_font"]
        cell.fill = styles["title_fill"]
        cell.alignment = styles["title_align"]
        ws.row_dimensions[title_row].height = _TITLE_ROW_HEIGHT
    if subtitle:
        title_row_count += 1
        sub_row = title_row_count
        ws.cell(row=sub_row, column=1, value=subtitle)
        if n_cols >= 2:
            ws.merge_cells(
                start_row=sub_row, start_column=1,
                end_row=sub_row, end_column=n_cols,
            )
        cell = ws.cell(row=sub_row, column=1)
        cell.font = styles["subtitle_font"]
        cell.fill = styles["subtitle_fill"]
        cell.alignment = styles["subtitle_align"]
        ws.row_dimensions[sub_row].height = _SUBTITLE_ROW_HEIGHT

    header_row = title_row_count + 1
    first_data_row = header_row + 1

    # ── 表头样式 ─────────────────────────────────────────────────────────
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["header_align"]
        cell.border = styles["border"]
    ws.row_dimensions[header_row].height = _HEADER_ROW_HEIGHT

    # ── 人员交替配色映射（仅在有 merge_groups 时生效）─────────────────
    # 为每个数据行偏移预计算应使用的背景 fill（按人员块交替）
    band_fill_map: dict[int, Any] = {}
    if merge_groups:
        band_fills = styles.get("band_fills", ())
        if band_fills:
            for group_idx, (start_off, end_off, _n_base) in enumerate(merge_groups):
                fill = band_fills[group_idx % len(band_fills)]
                for off in range(start_off, end_off + 1):
                    band_fill_map[off] = fill

    # ── 考勤结果条件配色准备 ────────────────────────────────────────────
    # attend_result_columns: 月度汇总场景，指定哪些列（0-based）是考勤结果列
    # attend_result_rows:    日历表场景，指定哪些行偏移（0-based）是考勤结果行
    # 两者都需要结合日期数据列（跳过基础信息列）来判断是否需要配色
    from openpyxl.styles import PatternFill as _PF
    _attend_fill_cache: dict[str, _PF] = {}

    def _get_attend_fill(color_hex: str) -> _PF:
        """按颜色值缓存 PatternFill，避免重复创建。"""
        if color_hex not in _attend_fill_cache:
            _attend_fill_cache[color_hex] = _PF(fill_type="solid", fgColor=color_hex)
        return _attend_fill_cache[color_hex]

    _ar_cols = attend_result_columns or set()
    _ar_rows = attend_result_rows or set()

    # ── 数据行样式 ───────────────────────────────────────────────────────
    for row_offset in range(n_rows):
        excel_row = first_data_row + row_offset
        row_fill = band_fill_map.get(row_offset)
        is_attend_row = row_offset in _ar_rows
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.font = styles["data_font"]
            cell.border = styles["border"]

            # 背景色优先级：考勤结果条件配色 > 人员交替配色 > 默认无色
            col_zero = col_idx - 1
            value = rows[row_offset][col_zero] if col_zero < len(rows[row_offset]) else None
            is_attend_cell = (col_zero in _ar_cols) or (is_attend_row and col_zero >= 4)
            attend_color = None
            if is_attend_cell and value not in (None, ""):
                attend_color = _match_attend_result_color(str(value))

            if attend_color:
                cell.fill = _get_attend_fill(attend_color)
            elif row_fill is not None:
                cell.fill = row_fill

            # 数据区统一居中（参考图风格），仅对长文本（>10 字符）的非首列左对齐避免拥挤
            if _is_numeric_value(value):
                cell.alignment = styles["align_right"]
            else:
                text = str(value) if value not in (None, "") else ""
                if _estimate_text_width(text) > 14 and col_idx > 1:
                    cell.alignment = styles["align_left"]
                else:
                    cell.alignment = styles["align_center"]
        ws.row_dimensions[excel_row].height = _DATA_ROW_HEIGHT

    # ── 列宽自适应 ───────────────────────────────────────────────────────
    for col_idx in range(1, n_cols + 1):
        header_text = str(headers[col_idx - 1])
        max_width = _estimate_text_width(header_text) + 2
        for row in rows:
            if col_idx - 1 < len(row):
                cell_value = row[col_idx - 1]
                if cell_value is None or cell_value == "":
                    continue
                w = _estimate_text_width(str(cell_value))
                if w > max_width:
                    max_width = w
        max_width = min(max(max_width + 1, _MIN_COL_WIDTH), _MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_width

    # ── 冻结窗格 ─────────────────────────────────────────────────────────
    # 冻结到首个数据行 + 第二列（保留标题/表头/姓名列常驻）
    freeze_col_letter = "B" if (freeze_first_col and n_cols >= 2) else "A"
    ws.freeze_panes = f"{freeze_col_letter}{first_data_row}"

    # ── 自动筛选器（覆盖表头到最后一行数据）──────────────────────────
    # 注意：合并单元格时不应用 auto_filter（会和合并冲突）
    if n_cols >= 1 and n_rows >= 1 and not merge_groups:
        ws.auto_filter.ref = (
            f"A{header_row}:{last_col_letter}{first_data_row + n_rows - 1}"
        )

    # ── 单元格纵向合并（仅指定的基础列）──────────────────────────
    # merge_groups: [(start_row_offset, end_row_offset, n_base_cols), ...]
    # 其中 row_offset 是基于数据区第 0 行的偏移
    if merge_groups:
        for start_offset, end_offset, n_base_cols in merge_groups:
            if end_offset <= start_offset:
                continue
            excel_start = first_data_row + start_offset
            excel_end = first_data_row + end_offset
            for col_idx in range(1, n_base_cols + 1):
                # 取首行的值，合并后只保留首行内容
                top_value = ws.cell(row=excel_start, column=col_idx).value
                ws.merge_cells(
                    start_row=excel_start, start_column=col_idx,
                    end_row=excel_end, end_column=col_idx,
                )
                top_cell = ws.cell(row=excel_start, column=col_idx)
                top_cell.value = top_value
                top_cell.alignment = styles["align_center"]
                top_cell.border = styles["border"]

    # ── 图片嵌入（下载 URL 列指向的图片，转 PNG，嵌入对应单元格）────
    if image_columns:
        _embed_images_in_columns(
            ws, headers, rows,
            image_column_names=image_columns,
            header_row=header_row,
            first_data_row=first_data_row,
            image_size=image_size,
        )


def write_excel(
    out_path: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    sheet_name: str = "考勤报表",
    title: str | None = None,
    subtitle: str | None = None,
    image_columns: list[str] | None = None,
    image_size: tuple[int, int] = (80, 120),
) -> None:
    """
    用 openpyxl 写一个钉钉风格的美化版 Excel。

    布局：
      [可选] 第 1 行：主标题（青绿底色，跨所有列）
      [可选] 第 2 行：副标题（浅蓝底色，跨所有列，常用于"报表生成时间"）
      表头行：浅黄底色 + 黑字加粗 + 居中 + 加高
      数据行：白底 + 文本居中 / 数字右对齐 + 灰色细边框

    交互：
      - 冻结首列 + 表头（含标题区域）
      - 自动筛选器（覆盖表头到最后一行）
      - 列宽自适应（中文字符按 2 宽度估算）

    参数:
      title:    可选主标题（如"月度汇总展示  统计日期：2026-01-01 至 2026-01-31"）
      subtitle: 可选副标题（如"报表生成时间：2026-01-20 15:46"）
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError(
            "缺少 openpyxl 依赖，请执行：pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # openpyxl sheet 名 ≤ 31 字符

    # 计算标题占用行数
    title_row_count = (1 if title else 0) + (1 if subtitle else 0)
    header_row = title_row_count + 1
    first_data_row = header_row + 1

    # 写入表头
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)

    # 写入数据
    for row_offset, row in enumerate(rows):
        excel_row = first_data_row + row_offset
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=excel_row, column=col_idx, value=_excel_safe(value))

    # 应用统一样式（含标题区）
    styles = _build_styles()
    _apply_sheet_styles(
        ws, headers, rows, styles,
        title=title, subtitle=subtitle,
        image_columns=image_columns, image_size=image_size,
    )

    out_abs = os.path.abspath(out_path)
    wb.save(out_abs)


def _excel_safe(value: Any) -> Any:
    """
    把 dict / list 等复杂类型序列化为 JSON 字符串，避免 openpyxl 写入失败。
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.strftime(DATETIME_FMT)
    try:
        return json.dumps(value, ensure_ascii=False)
    except (TypeError, ValueError):
        return str(value)


def write_excel_multi_sheets(
    out_path: str,
    sheets: list[dict],
) -> None:
    """
    用 openpyxl 写一个多 sheet 的 Excel 文件，每个 sheet 应用与 write_excel 相同的钉钉风格美化样式。

    每个 sheet 用一个 dict 描述：
      {
        "name":          str,             # sheet 标题
        "headers":       list[str],
        "rows":          list[list[Any]],
        "title":         str | None,      # 可选主标题（青绿底）
        "subtitle":      str | None,      # 可选副标题（浅蓝底）
        "merge_groups":  list[tuple[int,int,int]] | None,
                                          # 可选纵向合并配置（基础列）
                                          # 每项: (start_row_offset, end_row_offset, n_base_cols)
        "attend_result_columns": set[int] | None,
                                          # 可选：考勤结果列的 0-based 列索引集合（月度汇总场景）
        "attend_result_rows":    set[int] | None,
                                          # 可选：考勤结果行的 0-based 行偏移集合（日历表场景）
        "image_columns": list[str] | None,
                                          # 可选：哪些列名的 URL 要嵌入为图片
        "image_size":    tuple[int,int] | None,
                                          # 可选：嵌入图片像素尺寸 (width, height)，默认 (80,120)
      }
    """
    if not sheets:
        raise ValueError("sheets 不能为空")

    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError(
            "缺少 openpyxl 依赖，请执行：pip install openpyxl"
        ) from e

    wb = Workbook()
    # 删除默认 sheet，由 sheets 描述完全决定
    default_ws = wb.active
    wb.remove(default_ws)

    styles = _build_styles()

    for sheet_def in sheets:
        name = sheet_def.get("name") or "Sheet"
        headers = sheet_def.get("headers") or []
        rows = sheet_def.get("rows") or []
        title = sheet_def.get("title")
        subtitle = sheet_def.get("subtitle")
        merge_groups = sheet_def.get("merge_groups")
        attend_result_columns = sheet_def.get("attend_result_columns")
        attend_result_rows = sheet_def.get("attend_result_rows")
        image_columns = sheet_def.get("image_columns")
        image_size = sheet_def.get("image_size") or (80, 120)

        ws = wb.create_sheet(title=name[:31])

        title_row_count = (1 if title else 0) + (1 if subtitle else 0)
        header_row = title_row_count + 1
        first_data_row = header_row + 1

        # 写入表头
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col_idx, value=header)

        # 写入数据
        for row_offset, row in enumerate(rows):
            excel_row = first_data_row + row_offset
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=excel_row, column=col_idx, value=_excel_safe(value))

        # 应用统一样式
        if headers:
            _apply_sheet_styles(
                ws, headers, rows, styles,
                title=title, subtitle=subtitle,
                merge_groups=merge_groups,
                attend_result_columns=attend_result_columns,
                attend_result_rows=attend_result_rows,
                image_columns=image_columns,
                image_size=image_size,
            )

    out_abs = os.path.abspath(out_path)
    wb.save(out_abs)


# ─────────────────────────────────────────────────────────────────────────────
# 输出文件命名
# ─────────────────────────────────────────────────────────────────────────────

def build_output_filename(start: datetime, end: datetime, *, suffix: str = "") -> str:
    """
    生成 attendance_report_<startDate>_<endDate>[_suffix].xlsx 形式的文件名，
    落在当前工作目录。
    """
    base = f"attendance_report_{start.strftime(DATE_FMT)}_{end.strftime(DATE_FMT)}"
    if suffix:
        base = f"{base}_{suffix}"
    return f"{base}.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# 请假数据查询（query-leave）
# ─────────────────────────────────────────────────────────────────────────────

# 默认关注的 4 类假期
DEFAULT_LEAVE_NAMES: tuple[str, ...] = ("事假", "调休", "病假", "年假")


def _normalize_leave_date(raw: Any) -> str | None:
    """把 query-leave 返回的 date 字段（毫秒时间戳字符串/数字）归一化为 YYYY-MM-DD。"""
    if raw is None or raw == "":
        return None
    try:
        ts = int(str(raw).strip())
    except (TypeError, ValueError):
        # 也可能本来就是 YYYY-MM-DD
        s = str(raw).strip()
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            return s[:10]
        return None
    # 毫秒级
    if ts >= 1_000_000_000_000:
        ts = ts // 1000
    try:
        return datetime.fromtimestamp(ts).strftime(DATE_FMT)
    except (OSError, ValueError, OverflowError):
        return None


def query_leave_data(
    user_ids: list[str],
    start: datetime,
    end: datetime,
    leave_names: Iterable[str] = DEFAULT_LEAVE_NAMES,
    *,
    stats: "CallStats | None" = None,
) -> dict[str, dict[str, dict[str, float]]]:
    """
    分批分段调用 `dws attendance report query-leave`，聚合每个用户每天的假期数据。

    返回结构：
      {
        userId: {
          "YYYY-MM-DD": {
            "事假": 1.0,
            "调休": 0.5,
            ...
          },
          ...
        },
        ...
      }

    若同一 (userId, date, leaveName) 在多次返回中出现（理论不会），按 sum 累加。
    分批规则与 query-data 一致：≤ MAX_USERS_PER_BATCH 人/次、≤ MAX_DAYS_PER_SLICE 天/次。
    """
    result: dict[str, dict[str, dict[str, float]]] = {}
    if not user_ids:
        return result

    leave_names_list = [n for n in leave_names if n]
    if not leave_names_list:
        return result

    user_batches = chunk_users(user_ids)
    date_slices = slice_date_range(start, end)
    leave_arg = ",".join(leave_names_list)

    log(
        f"[leave] 查询 {len(user_ids)} 人 × {len(leave_names_list)} 类假期 × "
        f"{len(date_slices)} 个时间片"
    )

    for bi, batch in enumerate(user_batches, start=1):
        for si, dslice in enumerate(date_slices, start=1):
            log(
                f"[leave] [batch {bi}/{len(user_batches)}] "
                f"[slice {si}/{len(date_slices)}] users={len(batch)} "
                f"slice={dslice.label}"
            )
            try:
                payload = run_dws([
                    "attendance", "report", "query-leave",
                    "--users", ",".join(batch),
                    "--leave-names", leave_arg,
                    "--start", dslice.start_str,
                    "--end", dslice.end_str,
                ])
                if stats is not None:
                    stats.total_dws_calls += 1
            except DwsCallError as e:
                if stats is not None:
                    stats.total_dws_calls += 1
                    stats.failed_calls += 1
                if e.is_permission_error:
                    error("权限错误：当前账号无管理员权限，无法查询请假数据。")
                    raise SystemExit(2) from e
                if stats is not None:
                    stats.add_warning(f"[leave query failed] {dslice.label}: {e}")
                else:
                    warn(f"[leave query failed] {dslice.label}: {e}")
                continue

            records = extract_records(payload)
            for record in records:
                uid = _first_nonempty(record, ("userId", "userid", "user_id"))
                if uid is None:
                    continue
                uid_str = str(uid)
                leave_vals = record.get("leaveVals")
                if not isinstance(leave_vals, list):
                    continue
                user_bucket = result.setdefault(uid_str, {})
                for entry in leave_vals:
                    if not isinstance(entry, dict):
                        continue
                    date_str = _normalize_leave_date(entry.get("date"))
                    if not date_str:
                        continue
                    leave_name = entry.get("leaveName") or entry.get("name")
                    if not leave_name:
                        continue
                    leave_name = str(leave_name)
                    raw_value = entry.get("value", entry.get("data", 0))
                    try:
                        num = float(str(raw_value).strip())
                    except (TypeError, ValueError):
                        continue
                    day_bucket = user_bucket.setdefault(date_str, {})
                    day_bucket[leave_name] = day_bucket.get(leave_name, 0.0) + num

    return result


def build_vacation_filename(
    start: datetime | None,
    end: datetime | None,
    *,
    as_of: datetime | None = None,
) -> str:
    """
    生成 vacation_export_<startDate>_<endDate>.xlsx；
    无时间区间时退化为 vacation_export_<asOfDate>.xlsx（asOfDate 默认为今天）。
    """
    if start is not None and end is not None:
        return f"vacation_export_{start.strftime(DATE_FMT)}_{end.strftime(DATE_FMT)}.xlsx"
    snapshot_date = (as_of or datetime.now()).strftime(DATE_FMT)
    return f"vacation_export_{snapshot_date}.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# 通用报告骨架
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class CallStats:
    """记录一次脚本运行中的 dws 调用统计，用于最终摘要。"""
    user_batches: int = 0
    date_slices: int = 0
    total_dws_calls: int = 0
    failed_calls: int = 0
    warnings: list[str] = field(default_factory=list)

    def add_warning(self, msg: str) -> None:
        self.warnings.append(msg)
        warn(msg)


def print_summary(
    *,
    granularity_label: str,
    out_path: str,
    user_count: int,
    column_names: list[str],
    start: datetime,
    end: datetime,
    rows_count: int,
    stats: CallStats,
    extra_tail: str = "",
) -> None:
    """
    把最终摘要打到 stdout，供调用方（Agent / 终端用户）查看。
    格式与 SKILL.md 输出模板对齐。
    """
    abs_path = os.path.abspath(out_path)
    print("[完成] 考勤报表已导出")
    print()
    print(f"[文件] {abs_path}")
    print(f"[粒度] {granularity_label}")
    print(f"[用户] {user_count} 人（共 {stats.user_batches} 批）")
    print(f"[时间] {start.strftime(DATE_FMT)} ~ {end.strftime(DATE_FMT)}"
          f"（共 {stats.date_slices} 个时间片）")
    print(f"[调用] 共调用 dws 接口：{stats.total_dws_calls} 次"
          + (f"（其中 {stats.failed_calls} 次失败）" if stats.failed_calls else ""))
    print(f"[字段] {' / '.join(column_names) if column_names else '（默认）'}")
    print(f"[行数] {rows_count} 行")
    if stats.warnings:
        print()
        print("[警告]")
        for w in stats.warnings[:10]:
            print(f"  - {w}")
        if len(stats.warnings) > 10:
            print(f"  - ...（共 {len(stats.warnings)} 条警告）")
    if extra_tail:
        print()
        print(extra_tail)


# ─────────────────────────────────────────────────────────────────────────────
# 图片下载 + Excel 嵌入（detail 报表的"打卡图片"列专用）
# ─────────────────────────────────────────────────────────────────────────────

# 全局缓存：URL → 本地 PNG 文件路径，避免同一张图重复下载/转换
_IMAGE_CACHE_DIR = os.path.join(
    tempfile.gettempdir(), "dws_attendance_report_images"
)
_image_url_to_local: dict[str, str] = {}
# 下载/转换失败的 URL 黑名单，避免反复重试
_image_failed_urls: set[str] = set()


def _ensure_image_cache_dir() -> str:
    """确保图片缓存目录存在并返回路径。"""
    os.makedirs(_IMAGE_CACHE_DIR, exist_ok=True)
    return _IMAGE_CACHE_DIR


def _is_likely_url(value: Any) -> bool:
    """简单判断一个值是不是 http(s) URL。"""
    if not isinstance(value, str):
        return False
    s = value.strip()
    return s.startswith("http://") or s.startswith("https://")


def download_and_convert_image(
    url: str,
    *,
    timeout: int = 10,
) -> str | None:
    """
    下载图片 URL → PIL 转 PNG → 缓存到本地，返回本地 PNG 文件路径。

    特性：
      - 磁盘缓存（同一 URL 只下载一次）
      - 支持 webp/jpg/jpeg/png 等格式（PIL 自动识别）
      - 失败的 URL 加黑名单，避免反复重试
      - 失败返回 None，调用方应保留原 URL 文本

    依赖: requests + Pillow（PIL）
    """
    if url in _image_url_to_local:
        return _image_url_to_local[url]
    if url in _image_failed_urls:
        return None

    cache_dir = _ensure_image_cache_dir()
    url_hash = hashlib.md5(url.encode("utf-8")).hexdigest()
    local_path = os.path.join(cache_dir, f"{url_hash}.png")

    if os.path.exists(local_path) and os.path.getsize(local_path) > 0:
        _image_url_to_local[url] = local_path
        return local_path

    try:
        import requests
    except ImportError:
        warn("缺少 requests 依赖，无法下载图片，请执行: pip install requests")
        _image_failed_urls.add(url)
        return None
    try:
        from PIL import Image as PILImage
    except ImportError:
        warn("缺少 Pillow 依赖，无法转换图片格式，请执行: pip install Pillow")
        _image_failed_urls.add(url)
        return None

    try:
        resp = requests.get(url, timeout=timeout)
        resp.raise_for_status()
        raw_bytes = resp.content
        if not raw_bytes:
            raise ValueError("下载结果为空")
    except Exception as e:
        warn(f"[image] 下载失败: {url[:80]}... 原因: {e}")
        _image_failed_urls.add(url)
        return None

    try:
        from io import BytesIO
        with PILImage.open(BytesIO(raw_bytes)) as img:
            # webp 等可能是 RGBA / P 模式，统一转 RGB 再存 PNG
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGBA")
            img.save(local_path, format="PNG")
    except Exception as e:
        warn(f"[image] 转换失败: {url[:80]}... 原因: {e}")
        _image_failed_urls.add(url)
        return None

    _image_url_to_local[url] = local_path
    return local_path


def _set_image_hyperlink(ws, excel_row: int, col_idx: int, url: str) -> None:
    """把单元格设为可点击的超链接，文案显示"打卡图片"，避免直接暴露裸 URL。"""
    from openpyxl.styles import Font
    cell = ws.cell(row=excel_row, column=col_idx)
    cell.value = "打卡图片"
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")


def _replace_all_image_urls_with_hyperlinks(
    ws, headers: list[str], rows: list[list[Any]],
    image_column_names: list[str], first_data_row: int,
) -> None:
    """Pillow 不可用时的兜底：把所有图片列的 URL 替换为"打卡图片"超链接。"""
    name_to_col_idx: dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        if h in image_column_names and h not in name_to_col_idx:
            name_to_col_idx[h] = i
    for row_offset, row in enumerate(rows):
        excel_row = first_data_row + row_offset
        for col_idx in name_to_col_idx.values():
            if col_idx - 1 >= len(row):
                continue
            if _is_likely_url(row[col_idx - 1]):
                _set_image_hyperlink(ws, excel_row, col_idx, str(row[col_idx - 1]).strip())


def _embed_images_in_columns(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    *,
    image_column_names: list[str],
    header_row: int,
    first_data_row: int,
    image_size: tuple[int, int] = (80, 120),
) -> None:
    """
    把指定列里的 URL 替换为嵌入的图片：
      1. 找到 image_column_names 命中的列索引
      2. 遍历每行该列的值，若是 http(s) URL 则下载 + 转 PNG + add_image
      3. 同步调整列宽（与图片宽匹配）和行高（与图片高匹配）
      4. 下载/转换失败时将 URL 替换为可点击的"打卡图片"超链接，避免暴露裸 URL

    image_size: (width_px, height_px)，控制嵌入图片尺寸，默认 80×120 像素
    """
    if not image_column_names:
        return
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        warn(f"openpyxl 不完整，无法嵌入图片: {e}")
        return

    # openpyxl 的 Image 类内部依赖 Pillow（模块加载时检测），
    # 如果 Pillow 不可用，OpenpyxlImage() 会抛出：
    #   ImportError: You must install Pillow to fetch image objects
    # 这里提前检测，不可用时尝试自动安装，避免逐张图片重复报错。
    from openpyxl.drawing.image import PILImage as _openpyxl_pil_check
    if not _openpyxl_pil_check:
        warn(
            "[image] openpyxl 检测到 Pillow 未安装，尝试自动安装..."
        )
        import subprocess, sys
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "Pillow"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                timeout=60,
            )
            log("[image] Pillow 安装成功，重新加载 openpyxl.drawing.image...")
            # 安装后需要重新加载模块，让 openpyxl 重新检测 Pillow
            import importlib
            import openpyxl.drawing.image as _img_mod
            importlib.reload(_img_mod)
            from openpyxl.drawing.image import Image as OpenpyxlImage  # noqa: F811
            from openpyxl.drawing.image import PILImage as _recheck
            if not _recheck:
                warn(
                    "[image] Pillow 安装后 openpyxl 仍无法检测到，"
                    "图片将显示为可点击链接。请手动执行: pip install Pillow"
                )
                _replace_all_image_urls_with_hyperlinks(
                    ws, headers, rows, image_column_names, first_data_row,
                )
                return
        except Exception as install_err:
            warn(
                f"[image] Pillow 自动安装失败: {install_err}\n"
                "图片将显示为可点击链接。请手动执行: pip install Pillow"
            )
            _replace_all_image_urls_with_hyperlinks(
                ws, headers, rows, image_column_names, first_data_row,
            )
            return

    # 找到目标列索引（1-based）
    name_to_col_idx: dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        if h in image_column_names and h not in name_to_col_idx:
            name_to_col_idx[h] = i
    if not name_to_col_idx:
        return

    width_px, height_px = image_size
    # Excel 列宽单位 ≈ 字符数，1 字符 ≈ 7px；行高单位为点，1 点 ≈ 1.33px
    col_width = max(width_px / 7.0, 12.0)
    row_height = max(height_px * 0.78, 60.0)

    # 调列宽
    for col_idx in name_to_col_idx.values():
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 收集所有要嵌入的 (excel_row, col_idx, url)，统一处理
    embed_tasks: list[tuple[int, int, str]] = []
    for row_offset, row in enumerate(rows):
        excel_row = first_data_row + row_offset
        row_has_image = False
        for col_name, col_idx in name_to_col_idx.items():
            if col_idx - 1 >= len(row):
                continue
            value = row[col_idx - 1]
            if not _is_likely_url(value):
                continue
            url = str(value).strip()
            embed_tasks.append((excel_row, col_idx, url))
            row_has_image = True
        if row_has_image:
            ws.row_dimensions[excel_row].height = row_height

    if not embed_tasks:
        return

    log(f"[image] 准备嵌入 {len(embed_tasks)} 张图片到 Excel...")
    success_count = 0
    failed_count = 0

    for excel_row, col_idx, url in embed_tasks:
        local_path = download_and_convert_image(url)
        if not local_path:
            _set_image_hyperlink(ws, excel_row, col_idx, url)
            failed_count += 1
            continue
        try:
            img = OpenpyxlImage(local_path)
            img.width = width_px
            img.height = height_px
            anchor = f"{get_column_letter(col_idx)}{excel_row}"
            # 清空原 URL 单元格内容（图片浮在格子上，但保留文本会重叠）
            ws.cell(row=excel_row, column=col_idx, value="")
            ws.add_image(img, anchor)
            success_count += 1
        except Exception as e:
            warn(f"[image] 嵌入失败 ({excel_row},{col_idx}): {e}")
            _set_image_hyperlink(ws, excel_row, col_idx, url)
            failed_count += 1

    log(f"[image] 嵌入完成: 成功 {success_count} 张，失败 {failed_count} 张")
